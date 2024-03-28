import sys
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

class Site:
    """Represents an OTTA site

    attributes: acronym, description, otta_acronym
    """

    def __init__(self, acronym, description, ocac_acronym = None):
        self.acronym = acronym
        self.description = description
        self.ocac_acronym = ocac_acronym

class SiteGroup:
    """A group of OTTA sites

    Can get a site by acronym and return a list of acronyms in order
    """

    def __init__(self, otta_site_excel_path):
        self.map = {}
        wb: Workbook = load_workbook(filename=otta_site_excel_path)
        sheet = wb.active
        max_row = sheet.max_row
        for i in range(1, max_row + 1):
            acronym = sheet.cell(row=i, column=1).value
            # This indicates an empty row, probably at the end of the spreadsheet
            if not acronym:
                continue
            # skip header
            if (acronym == "OTTA Site"):
                continue
            if (self.has_acronym(acronym)):
                print("Duplicate OTTA site " + acronym)
                sys.exit(1)
            description = sheet.cell(row=i, column=2).value
            ocac_acronym = sheet.cell(row=i, column=3).value
            if (ocac_acronym and ocac_acronym.strip() == ""):
                ocac_acronym = None
            self.map[acronym] = Site(acronym, description, ocac_acronym)

    def ordered_acronyms(self):
        # acronym_list = []
        # for acronym in self.map.keys():
        #     acronym_list.append(acronym)
        # return acronym_list.sort()
        return sorted(self.map.keys())

    def has_acronym(self, acronym):
        return acronym in self.map.keys()

    def site_for_acronym(self, acronym):
        if (acronym in self.map):
            return self.map[acronym]
        raise KeyError("No site for acronym " + acronym)