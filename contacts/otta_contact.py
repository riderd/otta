import sys
from tabulate import tabulate
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

def clean(str):
   if not str:
       return ""
   return str.strip().replace("\"", "")

class Contact:
    """OTTA contact

    attributes: acronym, name, email, primary_contact
    """

    def __init__(self, acronym, name, email, primary_contact):
        self.acronym = acronym
        self.name = name
        self.email = email
        self.primary_contact = primary_contact

    def primary_yes_no(self):
        if self.primary_contact:
            return "Yes"
        return "No"

class ContactList:
    """List of OTTA contacts

    get list of primary contacts by acronym
    """


    def __init__(self, otta_contacts_excel_path, public_excluded_sites_filename=None):
        self.missing_otta_sites = False

        wb: Workbook = load_workbook(filename=otta_contacts_excel_path)
        sheet: Worksheet = wb.active
        max_row = sheet.max_row

        self.contacts = {}
        for i in range(1, max_row + 1):
            acronym = clean(sheet.cell(row=i, column=1).value)

            primary_contact = False;
            yes_no = sheet.cell(row=i, column=2).value
            if (yes_no.lower() == "yes"):
                primary_contact = True;

            name = clean(sheet.cell(row=i, column=3).value)
            email = clean(sheet.cell(row=i, column=4).value)

            contact = Contact(acronym, name, email, primary_contact)
            contact_list_for_acronym = []
            if acronym in self.contacts:
                contact_list_for_acronym = self.contacts[acronym]
            contact_list_for_acronym.append(contact)
            self.contacts[acronym] = contact_list_for_acronym

        self.excluded_public_groups = self._read_excluded_public_groups(public_excluded_sites_filename)

    def _read_excluded_public_groups(self, public_excluded_sites_filename):
        if not public_excluded_sites_filename:
            return None

        excluded_groups = []
        with open(public_excluded_sites_filename) as fp:
            Lines = fp.readlines()
            for line in Lines:
                excluded_groups.append(line.strip())
        return excluded_groups

    def missing_sites(self, sites):
        for acronym in self.contacts:
            if (not acronym in sites.ordered_acronyms()):
                return True
        return False

    def print_contacts_without_sites(self, sites):
        if not self.missing_sites(sites):
            return

        print("Some OTTA sites from contact(s) are missing from master list", file=sys.stderr)
        for acronym in self.contacts:
            if not acronym in sites.ordered_acronyms():
                names = []
                for contact in self.contacts[acronym]:
                    names.append(contact.name)
                print("%s: %s" % (acronym, names), file=sys.stderr)
                print()

    def _add_class(self, table_html:str):
        return table_html.replace("<table>", '<table class="otta-table">')

    def print_public_table(self, sites):
        table_data = []
        for acronym in sites.ordered_acronyms():
            num_primary_contacts_for_site = 0
            if acronym in self.excluded_public_groups:
                print(acronym + " excluded from public table", file=sys.stderr)
                continue
            site = sites.site_for_acronym(acronym)
            if not acronym in self.contacts:
                print("No contacts listed for OTTA site: " + acronym, file=sys.stderr)
                continue
            for contact in self.contacts[acronym]:
                if contact.primary_contact:
                    num_primary_contacts_for_site += 1
                    table_data.append([acronym, site.description, contact.name, contact.email, site.ocac_acronym])

            # maybe put out a warning for an OTTA site if there is no primary contact?
            if not num_primary_contacts_for_site:
                print("Couldn't find a primary contact for " + acronym, file=sys.stderr)
            if num_primary_contacts_for_site > 1:
                print("Published multiple primary contacts for " + acronym, file=sys.stderr)

        table_html = tabulate(table_data, headers=["OTTA Site", "Study Name", "Contact", "Email", "OCAC"], tablefmt="html")
        print(self._add_class(table_html))

    def print_member_table(self, sites):
        table_data = []
        for acronym in sites.ordered_acronyms():
            if not acronym in self.contacts:
                print("No contacts listed for OTTA site: " + acronym)
                continue
            for contact in self.contacts[acronym]:
                table_data.append([acronym, contact.primary_yes_no(), contact.name, contact.email])

        table_html = tabulate(table_data, headers=["OTTA Study Acronym", "Primary Contact", "Name", "Email"], tablefmt="html")
        print(self._add_class(table_html))
